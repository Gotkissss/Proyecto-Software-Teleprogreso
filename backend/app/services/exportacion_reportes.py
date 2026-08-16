"""
Generacion de archivos Excel para los reportes agregados.

Este modulo solo transforma los resultados ya calculados por reportes.py; no
consulta la base de datos ni vuelve a implementar reglas de agregacion.
"""

from datetime import timedelta
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.schemas.reporte import (
    ReporteAsistenciaResponse,
    ReporteProductividadResponse,
    ReporteTareasCompletadasResponse,
)

TIPOS_REPORTES_EXPORTABLES = (
    "asistencia",
    "tareas-completadas",
    "productividad",
)

MIME_XLSX = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

COLOR_TITULO = "1F4E78"
COLOR_ENCABEZADO = "5B9BD5"
COLOR_TOTAL = "D9EAF7"


def _ajustar_columnas(hoja) -> None:
    """Ajusta anchos con un limite para evitar columnas desproporcionadas."""
    for columna in hoja.columns:
        letra = get_column_letter(columna[0].column)
        ancho = max(
            (len(str(celda.value)) for celda in columna if celda.value is not None),
            default=0,
        )
        hoja.column_dimensions[letra].width = min(max(ancho + 2, 12), 40)


def _crear_libro(
    *,
    titulo: str,
    nombre_hoja: str,
    reporte,
    encabezados: tuple[str, ...],
    filas: list[tuple],
    fila_total: tuple,
    columnas_duracion: tuple[int, ...] = (),
    columnas_decimal: tuple[int, ...] = (),
) -> bytes:
    """Construye un libro con titulo, periodo, tabla y totales."""
    libro = Workbook()
    hoja = libro.active
    hoja.title = nombre_hoja
    ultima_columna = get_column_letter(len(encabezados))

    hoja.merge_cells(f"A1:{ultima_columna}1")
    hoja["A1"] = titulo
    hoja["A1"].font = Font(bold=True, color="FFFFFF", size=14)
    hoja["A1"].fill = PatternFill("solid", fgColor=COLOR_TITULO)
    hoja["A1"].alignment = Alignment(horizontal="center")

    hoja.merge_cells(f"A2:{ultima_columna}2")
    hoja["A2"] = (
        f"Periodo: {reporte.fecha_inicio.isoformat()} al "
        f"{reporte.fecha_fin.isoformat()}"
    )
    hoja["A2"].alignment = Alignment(horizontal="center")

    fila_encabezados = 4
    for columna, encabezado in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila_encabezados, column=columna, value=encabezado)
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = PatternFill("solid", fgColor=COLOR_ENCABEZADO)
        celda.alignment = Alignment(horizontal="center")

    primera_fila_datos = fila_encabezados + 1
    for fila in filas:
        hoja.append(fila)

    fila_fin_datos = fila_encabezados + len(filas)
    hoja.append(fila_total)
    fila_totales = fila_fin_datos + 1

    for celda in hoja[fila_totales]:
        celda.font = Font(bold=True)
        celda.fill = PatternFill("solid", fgColor=COLOR_TOTAL)

    for numero_columna in columnas_duracion:
        for numero_fila in range(primera_fila_datos, fila_totales + 1):
            hoja.cell(numero_fila, numero_columna).number_format = "[h]:mm"

    for numero_columna in columnas_decimal:
        for numero_fila in range(primera_fila_datos, fila_totales + 1):
            hoja.cell(numero_fila, numero_columna).number_format = "0.00"

    hoja.freeze_panes = f"A{primera_fila_datos}"
    hoja.auto_filter.ref = (
        f"A{fila_encabezados}:{ultima_columna}"
        f"{max(fila_encabezados, fila_fin_datos)}"
    )
    _ajustar_columnas(hoja)

    contenido = BytesIO()
    libro.save(contenido)
    return contenido.getvalue()


def _excel_asistencia(reporte: ReporteAsistenciaResponse) -> bytes:
    filas = [
        (
            item.id_empleado,
            item.nombre_empleado,
            item.jornadas,
            item.jornadas_abiertas,
            timedelta(minutes=item.minutos_trabajados),
            timedelta(minutes=item.minutos_pausa),
        )
        for item in reporte.items
    ]

    return _crear_libro(
        titulo="Reporte de asistencia",
        nombre_hoja="Asistencia",
        reporte=reporte,
        encabezados=(
            "ID empleado",
            "Empleado",
            "Jornadas",
            "Jornadas abiertas",
            "Horas trabajadas",
            "Horas de pausa",
        ),
        filas=filas,
        fila_total=(
            None,
            "TOTAL",
            reporte.total_jornadas,
            sum(item.jornadas_abiertas for item in reporte.items),
            timedelta(minutes=reporte.total_minutos_trabajados),
            timedelta(minutes=reporte.total_minutos_pausa),
        ),
        columnas_duracion=(5, 6),
    )


def _excel_tareas_completadas(
    reporte: ReporteTareasCompletadasResponse,
) -> bytes:
    filas = [
        (
            item.id_empleado,
            item.nombre_empleado,
            item.tareas_completadas,
        )
        for item in reporte.items
    ]

    return _crear_libro(
        titulo="Reporte de tareas completadas",
        nombre_hoja="Tareas completadas",
        reporte=reporte,
        encabezados=("ID técnico", "Técnico", "Tareas completadas"),
        filas=filas,
        fila_total=(None, "TOTAL", reporte.total_tareas_completadas),
    )


def _excel_productividad(reporte: ReporteProductividadResponse) -> bytes:
    filas = [
        (
            item.id_empleado,
            item.nombre_empleado,
            item.jornadas,
            timedelta(minutes=item.minutos_trabajados),
            item.tareas_completadas,
            item.tareas_por_hora,
        )
        for item in reporte.items
    ]

    return _crear_libro(
        titulo="Reporte de productividad",
        nombre_hoja="Productividad",
        reporte=reporte,
        encabezados=(
            "ID técnico",
            "Técnico",
            "Jornadas",
            "Horas trabajadas",
            "Tareas completadas",
            "Tareas por hora",
        ),
        filas=filas,
        fila_total=(
            None,
            "TOTAL",
            reporte.total_jornadas,
            timedelta(minutes=reporte.total_minutos_trabajados),
            reporte.total_tareas_completadas,
            reporte.tareas_por_hora,
        ),
        columnas_duracion=(4,),
        columnas_decimal=(6,),
    )


def generar_excel_reporte(tipo: str, reporte) -> bytes:
    """Genera el XLSX correspondiente al tipo de reporte solicitado."""
    generadores = {
        "asistencia": _excel_asistencia,
        "tareas-completadas": _excel_tareas_completadas,
        "productividad": _excel_productividad,
    }
    try:
        generador = generadores[tipo]
    except KeyError as error:
        raise ValueError(f"Tipo de reporte no soportado: {tipo}") from error

    return generador(reporte)


def nombre_archivo_reporte(tipo: str, fecha_inicio, fecha_fin) -> str:
    """Nombre estable y seguro para la descarga del reporte."""
    tipo_archivo = tipo.replace("-", "_")
    return (
        f"reporte_{tipo_archivo}_{fecha_inicio.isoformat()}_"
        f"{fecha_fin.isoformat()}.xlsx"
    )

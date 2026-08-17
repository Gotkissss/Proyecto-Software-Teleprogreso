"""
Pruebas de la exportacion XLSX de los reportes agregados.
"""

from datetime import date, timedelta
from io import BytesIO
from types import SimpleNamespace
from unittest.mock import AsyncMock, MagicMock, patch

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook

from app.core.deps import require_admin_supervisor_gerente
from app.routers.reportes import exportar_reporte, router
from app.schemas.reporte import (
    AsistenciaEmpleado,
    ProductividadEmpleado,
    ReporteAsistenciaResponse,
    ReporteProductividadResponse,
    ReporteTareasCompletadasResponse,
    TareasCompletadasEmpleado,
)
from app.services.exportacion_reportes import (
    MIME_XLSX,
    generar_excel_reporte,
    nombre_archivo_reporte,
)

FECHA_INICIO = date(2026, 8, 1)
FECHA_FIN = date(2026, 8, 15)


def test_ruta_exportacion_permite_roles_de_supervision():
    ruta = next(
        ruta
        for ruta in router.routes
        if ruta.path == "/reportes/{tipo}/exportar"
    )
    dependencias = {dependencia.call for dependencia in ruta.dependant.dependencies}

    assert require_admin_supervisor_gerente in dependencias


@pytest.mark.asyncio
async def test_permiso_de_exportacion_acepta_supervisor_y_rechaza_tecnico():
    supervisor = SimpleNamespace(rol="supervisor")
    tecnico = SimpleNamespace(rol="tecnico")

    assert await require_admin_supervisor_gerente(supervisor) is supervisor

    with pytest.raises(HTTPException) as error:
        await require_admin_supervisor_gerente(tecnico)

    assert error.value.status_code == 403


def _reporte_asistencia():
    return ReporteAsistenciaResponse(
        fecha_inicio=FECHA_INICIO,
        fecha_fin=FECHA_FIN,
        total_empleados=1,
        total_jornadas=2,
        total_minutos_trabajados=720,
        total_minutos_pausa=60,
        total_horas_trabajadas="12:00",
        total_horas_pausa="01:00",
        items=[
            AsistenciaEmpleado(
                id_empleado=7,
                nombre_empleado="Ana Lopez",
                jornadas=2,
                jornadas_abiertas=0,
                minutos_trabajados=720,
                minutos_pausa=60,
                horas_trabajadas="12:00",
                horas_pausa="01:00",
            )
        ],
    )


def _reporte_tareas():
    return ReporteTareasCompletadasResponse(
        fecha_inicio=FECHA_INICIO,
        fecha_fin=FECHA_FIN,
        total_empleados=1,
        total_tareas_completadas=3,
        items=[
            TareasCompletadasEmpleado(
                id_empleado=7,
                nombre_empleado="Ana Lopez",
                tareas_completadas=3,
            )
        ],
    )


def _reporte_productividad():
    return ReporteProductividadResponse(
        fecha_inicio=FECHA_INICIO,
        fecha_fin=FECHA_FIN,
        total_empleados=1,
        total_jornadas=2,
        total_minutos_trabajados=720,
        total_horas_trabajadas="12:00",
        total_tareas_completadas=3,
        tareas_por_hora=0.25,
        items=[
            ProductividadEmpleado(
                id_empleado=7,
                nombre_empleado="Ana Lopez",
                jornadas=2,
                minutos_trabajados=720,
                horas_trabajadas="12:00",
                tareas_completadas=3,
                tareas_por_hora=0.25,
            )
        ],
    )


def test_excel_asistencia_contiene_duraciones_y_totales():
    contenido = generar_excel_reporte("asistencia", _reporte_asistencia())
    hoja = load_workbook(BytesIO(contenido))["Asistencia"]

    assert hoja["A1"].value == "Reporte de asistencia"
    assert hoja["A2"].value == "Periodo: 2026-08-01 al 2026-08-15"
    assert hoja["A4"].value == "ID empleado"
    assert hoja["B5"].value == "Ana Lopez"
    assert hoja["E5"].value == timedelta(hours=12)
    assert hoja["E5"].number_format == "[h]:mm"
    assert hoja["B6"].value == "TOTAL"
    assert hoja["E6"].value == timedelta(hours=12)
    assert hoja.freeze_panes == "A5"
    assert hoja.auto_filter.ref == "A4:F5"


def test_excel_tareas_completadas_contiene_conteos():
    contenido = generar_excel_reporte(
        "tareas-completadas",
        _reporte_tareas(),
    )
    hoja = load_workbook(BytesIO(contenido))["Tareas completadas"]

    assert hoja["A4"].value == "ID técnico"
    assert hoja["B5"].value == "Ana Lopez"
    assert hoja["C5"].value == 3
    assert hoja["B6"].value == "TOTAL"
    assert hoja["C6"].value == 3


def test_excel_productividad_contiene_formula_calculada():
    contenido = generar_excel_reporte(
        "productividad",
        _reporte_productividad(),
    )
    hoja = load_workbook(BytesIO(contenido))["Productividad"]

    assert hoja["D5"].value == timedelta(hours=12)
    assert hoja["E5"].value == 3
    assert hoja["F5"].value == 0.25
    assert hoja["F5"].number_format == "0.00"
    assert hoja["F6"].value == 0.25


def test_generador_rechaza_tipo_desconocido():
    with pytest.raises(ValueError):
        generar_excel_reporte("desconocido", _reporte_asistencia())


def test_nombre_de_archivo_es_estable_y_seguro():
    assert nombre_archivo_reporte(
        "tareas-completadas",
        FECHA_INICIO,
        FECHA_FIN,
    ) == "reporte_tareas_completadas_2026-08-01_2026-08-15.xlsx"


@pytest.mark.asyncio
async def test_endpoint_exporta_asistencia_con_filtro_de_empleado():
    db = MagicMock()
    gerente = SimpleNamespace(id_empleado=4, rol="gerente")
    reporte = _reporte_asistencia()

    with (
        patch(
            "app.routers.reportes.obtener_reporte_asistencia",
            new=AsyncMock(return_value=reporte),
        ) as servicio,
        patch(
            "app.routers.reportes.generar_excel_reporte",
            return_value=b"contenido-xlsx",
        ) as generador,
    ):
        respuesta = await exportar_reporte(
            db,
            gerente,
            "asistencia",
            FECHA_INICIO,
            FECHA_FIN,
            empleado=7,
        )

    servicio.assert_awaited_once_with(
        db,
        FECHA_INICIO,
        FECHA_FIN,
        id_empleado=7,
    )
    generador.assert_called_once_with("asistencia", reporte)
    assert respuesta.body == b"contenido-xlsx"
    assert respuesta.media_type == MIME_XLSX
    assert "reporte_asistencia_2026-08-01_2026-08-15.xlsx" in (
        respuesta.headers["content-disposition"]
    )


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("tipo", "nombre_servicio"),
    [
        ("tareas-completadas", "obtener_reporte_tareas_completadas"),
        ("productividad", "obtener_reporte_productividad"),
    ],
)
async def test_endpoint_exporta_reportes_con_filtro_de_tecnico(
    tipo,
    nombre_servicio,
):
    db = MagicMock()
    gerente = SimpleNamespace(id_empleado=4, rol="gerente")
    reporte = (
        _reporte_tareas()
        if tipo == "tareas-completadas"
        else _reporte_productividad()
    )

    with (
        patch(
            f"app.routers.reportes.{nombre_servicio}",
            new=AsyncMock(return_value=reporte),
        ) as servicio,
        patch(
            "app.routers.reportes.generar_excel_reporte",
            return_value=b"contenido-xlsx",
        ),
    ):
        respuesta = await exportar_reporte(
            db,
            gerente,
            tipo,
            FECHA_INICIO,
            FECHA_FIN,
            tecnico=7,
        )

    servicio.assert_awaited_once_with(
        db,
        FECHA_INICIO,
        FECHA_FIN,
        id_tecnico=7,
    )
    assert respuesta.body == b"contenido-xlsx"


@pytest.mark.asyncio
async def test_endpoint_rechaza_tipo_desconocido():
    with pytest.raises(HTTPException) as error:
        await exportar_reporte(
            MagicMock(),
            SimpleNamespace(rol="gerente"),
            "desconocido",
            FECHA_INICIO,
            FECHA_FIN,
        )

    assert error.value.status_code == 400


@pytest.mark.asyncio
@pytest.mark.parametrize(
    ("tipo", "empleado", "tecnico"),
    [
        ("asistencia", None, 7),
        ("productividad", 7, None),
    ],
)
async def test_endpoint_rechaza_filtro_incompatible(tipo, empleado, tecnico):
    with pytest.raises(HTTPException) as error:
        await exportar_reporte(
            MagicMock(),
            SimpleNamespace(rol="gerente"),
            tipo,
            FECHA_INICIO,
            FECHA_FIN,
            empleado=empleado,
            tecnico=tecnico,
        )

    assert error.value.status_code == 400
